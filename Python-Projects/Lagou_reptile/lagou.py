import requests
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Alignment
from bs4 import BeautifulSoup
import time

headers = {
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'Host': 'www.lagou.com',
    'Origin': 'https://www.lagou.com',
    'Referer': 'https://www.lagou.com/jobs/list_python/p-city_252?px=default',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.25 Safari/537.36 Core/1.70.3754.400 QQBrowser/10.5.4034.400'
}
s = requests.Session()


def get_pageNum(search_name):
    # 252是那个网站关于成都的编号
    res = s.get(f'https://www.lagou.com/jobs/list_{search_name}/p-city_252?px=default#filterBox', headers=headers)
    # 接下里获取文章页数
    soup = BeautifulSoup(res.text, 'lxml')
    page_num = int(soup.find('span', class_='span totalNum').string)
    return page_num


def get_info(url, page, search_name):
    data = {
        'first': 'false',
        'pn': page,
        'kd': search_name}
    cookie = s.cookies
    res = s.post(url, headers=headers, cookies=cookie, data=data)
    json_data = res.json()
    list_con = json_data['content']['positionResult']['result']
    info_list = []
    for i in list_con:
        info = []
        info.append(i.get('positionName', '暂暂无'))
        info.append(i.get('companyShortName', '暂无'))
        info.append(i.get('district', '暂无'))
        info.append(i.get('companyFullName', '暂无'))
        info.append(i.get('industryField', '暂无'))
        info.append(i.get('companySize', '暂无'))
        info.append(i.get('education', '暂无'))
        info.append(i.get('workYear', '暂无'))
        info.append(i.get('salary', '暂无'))
        info_list.append(info)
    return info_list


def main():
    print('当前时间:', time.strftime('%Y-%m-%d', time.localtime(time.time())), time.strftime('%H:%M:%S', time.localtime()))
    search_name = 'python'
    cityName = quote('成都')  # 查询地点 但其实只能查成都 有个参数除非你自己去查252那个
    page_num = get_pageNum(search_name)  # 这个可以随便修改
    wb = Workbook()  # 打开 excel 工作簿
    ws1 = wb.active  # 获取第一个sheet
    ws1.title = search_name
    ws1.merge_cells('A1:I1')  # 合并单元格
    ws1.cell(1, 1).value = 'Python职位信息'  # 合并的单元格内容
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.append(
        ['positionName', 'companyShortName', 'district', 'companyFullName', 'industryField', 'companySize', 'education',
         'workYear', 'salary'])
    print(f'本次搜索结果共有{page_num}页')
    url = f'https://www.lagou.com/jobs/positionAjax.json?city={cityName}&needAddtionalResult=false'
    for i in range(page_num - 28):
        list = get_info(url, i + 1, 'python')
        for i in list:
            ws1.append(i)
        time.sleep(2)
    ws1.column_dimensions['A'].width = 36
    ws1.column_dimensions['B'].width = 35
    ws1.column_dimensions['C'].width = 8.78
    ws1.column_dimensions['D'].width = 44
    ws1.column_dimensions['E'].width = 23
    ws1.column_dimensions['F'].width = 12
    ws1.column_dimensions['G'].width = 9.78
    ws1.column_dimensions['H'].width = 10.89
    ws1.column_dimensions['I'].width = 8.11
    wb.save(f'{search_name}职位信息.xlsx')


if __name__ == '__main__':
    main()
